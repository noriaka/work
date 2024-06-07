import os
from typing import Dict, Any, Union, Tuple
from bs4 import BeautifulSoup
from enum import Enum
import difflib
import openpyxl
import markdown


class Excel_type(Enum):
    event = 1
    action = 2
    function = 3


class Eca_func_type(Enum):
    none = 1
    nil = 2
    y3 = 3
    other = 4


class Eca_func:
    def __init__(self, key: str, lua_text: str, description: str,
                 func_type: Eca_func_type, param_num: int):
        self.key = key
        self.lua_text = lua_text
        self.description = description
        self.func_type = func_type
        self.param_num = param_num


class Lua_func:
    def __init__(self, name: str, func: str, param_num: int, description: str):
        self.name = name
        self.func = func
        self.param_num = param_num
        self.description = description


def excel_process(excel_path: str, excel_type: Excel_type) -> dict[Any, Eca_func]:
    """
    处理表返回eca字典
    :param
        excel_path: 表文件路径
        excel_type: 表的类型
    :return: eca字典 # {'NONE_ABILITY': {'类型': type, 'lua文本': nil}}
    """
    # 读取 Excel 文件
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook['触发器']

    column_names_cn = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    # column_names_en = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    lua_map = {}
    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, values_only=True):
        if row[0] is not None:
            if excel_type == Excel_type.event:
                key = None
                lua_text = None
                description = None
            elif excel_type == Excel_type.action:
                key = str(row[1])
                lua_text = str(row[10])
                description = str(row[15])
            elif excel_type == Excel_type.function:
                key = str(row[1])
                lua_text = str(row[9])
                description = str(row[10])
            else:
                key = None
                lua_text = None
                description = None

            # 判断lua函数类型
            if lua_text is None:
                func_type = Eca_func_type.none
            elif lua_text == 'nil':
                func_type = Eca_func_type.nil
            elif lua_text[:2] == 'y3':
                func_type = Eca_func_type.y3
            else:
                func_type = Eca_func_type.other

            # 计算lua函数参数个数
            if lua_text == Eca_func_type.none or lua_text == Eca_func_type.nil:
                param_num = 0
            else:
                l_idx = str(lua_text).rfind('(')
                r_idx = str(lua_text).rfind(')')
                if l_idx + 1 == r_idx:
                    param_num = 0
                else:
                    param_num = len(str(lua_text)[l_idx + 1:r_idx].strip().split(','))

            eca_func = Eca_func(key, lua_text, description, func_type, param_num)
            lua_map[key] = eca_func
        else:
            break
    return lua_map


def markdown_process(markdown_path: str) -> tuple[str, str, dict[str, Lua_func]]:
    """
    解析 Markdown 文件, 提取 ## 标题和 ```lua ... ``` 代码块中的内容。

    :param
        markdown_path: Markdown 文件路径
    :return: 包含标题和代码块内容以及函数参数个数的字典
    """
    result = {}

    with open(markdown_path, 'r', encoding='utf-8') as file:
        md_content = file.read()

    # 将 Markdown 转换为 HTML
    html_content = markdown.markdown(md_content)
    soup = BeautifulSoup(html_content, 'html.parser')
    api_type_en = soup.find('h1').text
    api_type = soup.find('h1').find_next_sibling().text
    # 获取所有 h2 标签的值
    func_name_h2 = soup.find_all('h2')
    for name_label in func_name_h2:
        name = name_label.text
        func_label = name_label.find_next_sibling()
        description_label = func_label.find_next_sibling()
        func = func_label.text.replace('lua', '').strip()
        l_idx = func.rfind('(')
        r_idx = func.rfind(')')
        if l_idx + 1 == r_idx:
            param_num = 0
        else:
            param_num = len(func[l_idx + 1:r_idx].strip().split(','))
        if description_label is not None:
            description = description_label.text
        else:
            description = 'None'
        lua_func = Lua_func(name, func, param_num, description)
        result[name] = lua_func
    return api_type, api_type_en, result


def lua_code_matching(eca_data: Eca_func, des2key_dict: dict[str, str], description_list: list,
                      lua_func_dict: dict[str, Lua_func]) -> str:
    """
    对表中的lua文本和lua函数进行匹配
    :param eca_data: 表中的eca对象
    :param des2key_dict: 描述和key的映射字典
    :param description_list: 描述列表
    :param lua_func_dict: lua函数字典
    :return: 匹配上的lua函数
    """
    if eca_data.func_type == Eca_func_type.none:
        result = 'None'
    elif eca_data.func_type == Eca_func_type.nil:
        result = 'nil'
    else:
        target = eca_data.description.split('-')[-1].strip()
        closest_match = difflib.get_close_matches(target, description_list, cutoff=0.5)
        if len(closest_match) == 0:
            result = '待定'
        elif len(closest_match) == 1:
            result = des2key_dict[str(closest_match[0])]
        else:
            # 如果参数都不匹配则默认返回第一个
            result = des2key_dict[str(closest_match[0])]

            # 如果相近的有两个以上则判断参数个数最符合的
            for i in range(len(closest_match)):
                key = des2key_dict[str(closest_match[i])]
                if eca_data.param_num == lua_func_dict[key].param_num:
                    result = key
                    break
    return result


def multi_markdown_process(root_dir):
    file_paths = []
    for dir_path, dir_names, file_names in os.walk(root_dir):
        for file_name in file_names:
            if file_name.endswith('.md'):
                file_path = os.path.join(dir_path, file_name)
                file_paths.append(file_path)
    lua_func_dict, des_dict, des2key_dict, api_type_dict = {}, {}, {}, {}
    for file_path in file_paths:
        api_type, api_type_en, res = markdown_process(file_path)
        api_type_dict[api_type] = api_type_en
        lua_func_dict[api_type] = res
        des_list = []
        des2key_dict_temp = {}
        for func_name, func_data in res.items():
            des_list.append(func_data.description)
            des2key_dict_temp[func_data.description] = func_name
        des_dict[api_type] = des_list
        des2key_dict[api_type] = des2key_dict_temp

    return lua_func_dict, des_dict, des2key_dict, api_type_dict


def generate_excel():
    lua_map = excel_process('04触发器函数表.xlsx', Excel_type.function)
    lua_func_dict, des_dict, des2key_dict, api_type_dict = multi_markdown_process('API')
    # with open('test.txt', 'w') as file:
    #     for k, v in lua_map.items():
    #         if v.description is not None:
    #             api_type = v.description.split('-')[0].strip()
    #             if api_type in lua_func_dict.keys():
    #                 lua_func = lua_code_matching(v, des2key_dict[api_type], des_dict[api_type], lua_func_dict[api_type])
    #             else:
    #                 lua_func = '待定'
    #         else:
    #             lua_func = 'None'
    #         file.write(f'{v.lua_text}\t\t{lua_func}\n')
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'lua函数映射表'
    sheet.cell(row=1, column=1, value='lua文本')
    sheet.cell(row=1, column=2, value='lua文本备注')
    sheet.cell(row=1, column=3, value='lua函数')
    sheet.cell(row=1, column=4, value='lua函数描述')
    cnt = 2
    for k, v in lua_map.items():
        lua_func = 'None'
        lua_func_description = 'None'
        if v.description is not None:
            api_type = v.description.split('-')[0].strip()
            if api_type in lua_func_dict.keys():
                key = lua_code_matching(v, des2key_dict[api_type], des_dict[api_type], lua_func_dict[api_type])
                lua_func = api_type_dict[api_type] + '-' + key
                if key != 'nil' and key != 'None' and key != '待定':
                    lua_func_description = lua_func_dict[api_type][key].description
            else:
                lua_func = '待定'
                lua_func_description = '待定'
        sheet.cell(row=cnt, column=1, value=v.lua_text)
        sheet.cell(row=cnt, column=2, value=v.description)
        sheet.cell(row=cnt, column=3, value=lua_func)
        sheet.cell(row=cnt, column=4, value=lua_func_description)
        cnt = cnt + 1
    workbook.save('lua函数映射表.xlsx')


class Lua_func_mapping_data:
    def __init__(self, lua_id:int, lua_text:str, lua_text_memo:str, lua_func:str, lun_func_memo:str, lua_func_unknown:bool):
        self.lua_id = lua_id
        self.lua_text = lua_text
        self.lua_text_memo = lua_text_memo
        self.lua_func = lua_func
        self.lun_func_memo = lun_func_memo
        self.lua_func_unknown = lua_func_unknown


def lua_func_mapping_table_processing(name):
    # 读取 Excel 文件
    workbook = openpyxl.load_workbook(f'{name}.xlsx')
    sheet = workbook[name]

    # column_names = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    lua_func_mapping_table_dict = {}
    lua_id = 0
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        if row[0] is not None:
            lua_id = lua_id + 1
            lua_text = str(row[0])
            lua_text_memo = str(row[1])
            lua_func = str(row[2])
            lua_func_memo = str(row[3])
            if lua_func == '待定':
                lua_func_unknown = True
            else:
                lua_func_unknown = False
            lua_func_mapping_data = Lua_func_mapping_data(lua_id, lua_text, lua_text_memo, lua_func, lua_func_memo, lua_func_unknown)
            lua_func_mapping_table_dict[lua_id] = lua_func_mapping_data
    return lua_func_mapping_table_dict


if __name__ == "__main__":
    # lua_func_mapping_table_dict = lua_func_mapping_table_processing('lua函数映射表')
    # print(len(lua_func_mapping_table_dict.keys()))
    # print(lua_func_mapping_table_dict[870].lua_text_memo)

    # excel_action = excel_process(excel_path='03触发器动作表.xlsx', excel_type=Excel_type.action)
    # workbook = openpyxl.Workbook()
    # sheet = workbook.active
    # sheet.title = '触发器动作映射lua函数表'
    # sheet.cell(row=1, column=1, value='lua文本')
    # sheet.cell(row=1, column=2, value='lua文本备注')
    # sheet.cell(row=1, column=3, value='lua函数')
    # sheet.cell(row=1, column=4, value='lua函数描述')
    # cnt = 2
    # for k, v in excel_action.items():
    #     sheet.cell(row=cnt, column=1, value=v.lua_text)
    #     sheet.cell(row=cnt, column=2, value=v.description)
    #     cnt = cnt + 1
    # workbook.save('触发器动作映射lua函数表.xlsx')

    # lua_map = excel_process(excel_path='03触发器动作表.xlsx', excel_type=Excel_type.action)
    # lua_func_dict, des_dict, des2key_dict, api_type_dict = multi_markdown_process('API')
    # workbook = openpyxl.load_workbook('触发器动作映射lua函数表.xlsx')
    # sheet = workbook.active
    # cnt = 2
    # for k, v in lua_map.items():
    #     lua_func = 'None'
    #     lua_func_description = 'None'
    #     if v.description is not None:
    #         api_type = v.description.split('-')[0].strip()
    #         if api_type in lua_func_dict.keys():
    #             key = lua_code_matching(v, des2key_dict[api_type], des_dict[api_type], lua_func_dict[api_type])
    #             lua_func = api_type_dict[api_type] + '-' + key
    #             if key != 'nil' and key != 'None' and key != '待定':
    #                 lua_func_description = lua_func_dict[api_type][key].description
    #         else:
    #             lua_func = '待定'
    #             lua_func_description = '待定'
    #
    #     if sheet.cell(row=cnt, column=3).value is None:
    #         sheet.cell(row=cnt, column=3, value=lua_func)
    #         sheet.cell(row=cnt, column=4, value=lua_func_description)
    #     cnt = cnt + 1
    # workbook.save('触发器动作映射lua函数表.xlsx')
    i = [1, 1, 0, 1]
    t = i[-3:] + i[:-3]
    print(t)
